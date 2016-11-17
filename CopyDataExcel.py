#This Script just get data for an excel and copy it to another excel if there is a match in some columns.
# FileTOCopy has a code but needs more data which is stored in FiletoCopy.
# this Script gets the code, search for it and copy data from FileToCopy to ModifyFIle


import openpyxl, os
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog
import ctypes

#msg box function
def msgBox(title, text, style):
    ctypes.windll.user32.MessageBoxW(0, text, title, style)


root = tk.Tk()
root.withdraw()


msgBox('File TO copy from', 'Choose file to copy data from', 1)
File_Copyfrom = filedialog.askopenfilename(title='Select File to Copy from')
print(File_Copyfrom)

root = tk.Tk()
root.withdraw()
msgBox('File to Modify', 'Choose file to modify', 1)

File_ToModify = filedialog.askopenfilename(title='Select File to Modify')
print(File_ToModify)

ColumToWrite = input('letter of column to write in FileToModify...')

KeyInCopy = input('number of column where the key is in FileToCopy...')
KeyInModify = input('number of column where the key is in FileTomodify...')


wbModif = openpyxl.load_workbook(File_ToModify)
wsModif = wbModif.active

wbToCopy = openpyxl.load_workbook(File_Copyfrom)
wsTocopy = wbToCopy.active

for modifRow in wsModif.iter_rows(row_offset=1):
	Code = str(modifRow[int(KeyInModify)].value)
	for copyRow in wsTocopy.iter_rows(row_offset=1):
		if str(copyRow[int(KeyInCopy)].value) == Code:
			CodRef = copyRow[0].value
			RowToWrite = str(modifRow[1].row)
			wsModif[ColumToWrite + RowToWrite] = CodRef
			break
wbModif.save(File_ToModify)
